import os
import argparse
import glob as gb
import re
from urllib.parse import unquote

import openpyxl


def prepare_header(new_book):
    header = ['Title', 'Location', 'Google URL', 'Google Job URL', 'Job', 'Company', 'Location', 'Timestamp', 'Type',
              'Salary', 'Description', 'Typical Salaries', 'Apply Link', 'Apply Link_link', 'Rank', 'Screen Capture']
    for col_num in range(1, 17):
        new_book.cell(row=1, column=col_num).value = header[col_num-1]

    return new_book


def prepare_sheet(records, destination_file):
    wb = openpyxl.Workbook()
    new_book = prepare_header(wb.active)
    row_no = 2

    for record in records:
        col_no = 1
        for key, value in record.items():
            new_book.cell(row=row_no, column=col_no).value = value
            col_no += 1

        row_no += 1

    print('Saving To File::', destination_file)
    wb.save(destination_file)


def load_primary_sheet(filename):
    wb = openpyxl.load_workbook(filename)
    stage1 = wb.active
    primary_records = {}

    first_row = True
    for row in stage1.rows:
        if first_row:
            first_row = False
            continue

        job_link_url = row[4].value
        job_link = job_link_url.split('#[!opt!]{"user_agent')[0]
        if job_link not in primary_records:
            primary_records[job_link] = [x.value for x in row]

    return primary_records


def load_secondary_sheet(file_key, path):
    other_files = gb.glob(f'{path}/{file_key}-*.xlsx')
    secondary_records = {}

    for other_file in other_files:
        print('Processing File From Stage 2 ::::', other_file)
        wb = openpyxl.load_workbook(other_file)
        stage2 = wb.active

        first_row = True
        for row in stage2.rows:
            if first_row:
                first_row = False
                continue

            raw_row = [x.value for x in row]
            if row[0].value in secondary_records:
                secondary_records[row[0].value].append(raw_row)
            else:
                secondary_records[row[0].value] = [raw_row]

    return secondary_records


def merge_records(primary_records, secondary_records):
    records = []
    for key, primary_value in primary_records.items():
        record = {}
        record['Title'] = unquote(''.join(re.findall(r'q=(.*)\+', primary_value[0])))
        record['Location'] = unquote(''.join(re.findall(r'\+(.*?)&', primary_value[0])))
        record['Google URL'] = primary_value[0]

        if key in secondary_records:
            for sc_record in secondary_records[key]:
                new_record = record.copy()
                new_record['Google Job URL'] = sc_record[0]
                new_record['Job'] = sc_record[1]
                new_record['Company'] = sc_record[2]
                new_record['Location_2'] = sc_record[3]
                new_record['Timestamp'] = sc_record[4]
                new_record['Type'] = sc_record[5]
                new_record['Salary'] = sc_record[6]
                new_record['Description'] = sc_record[7]
                new_record['Typical Salaries'] = sc_record[8]
                new_record['Apply Link'] = sc_record[9]
                new_record['Apply Link_link'] = sc_record[10]
                new_record['Rank'] = primary_value[5]
                new_record['Screen Capture'] = sc_record[12]
                records.append(new_record)
        else:
            record['Google Job URL'] = key
            record['Job'] = 'JOB EXPIRED'
            record['Company'] = ''
            record['Location_2'] = ''
            record['Timestamp'] = ''
            record['Type'] = ''
            record['Salary'] = ''
            record['Description'] = ''
            record['Typical Salaries'] = ''
            record['Apply Link'] =''
            record['Apply Link_link'] = ''
            record['Rank'] = primary_value[5]
            record['Screen Capture'] = ''
            records.append(record)

    return records


def parse_arguments():
    parser = argparse.ArgumentParser(description='Process command line arguments.')
    parser.add_argument('-stage1', type=dir_path)
    parser.add_argument('-stage2', type=dir_path)
    return parser.parse_args()


def dir_path(path):
    if os.path.isdir(path):
        return path

    raise argparse.ArgumentTypeError(f"readable_dir:{path} is not a valid path")


def evaluate(primary_filename, key, stage2_path, output_filename):
    primary_records = load_primary_sheet(primary_filename)
    secondary_records = load_secondary_sheet(key, stage2_path)
    records = merge_records(primary_records, secondary_records)
    prepare_sheet(records, output_filename)


def main():
    parsed_args = parse_arguments()
    stage1_filenames = gb.glob(f'{parsed_args.stage1}/*.xlsx')

    count = 1
    for primary_filename in stage1_filenames:
        file_name = primary_filename.split('/')[-1]
        print(f'Processing File {count} From Stage 1 ::::', file_name)
        key = file_name.replace('-m.xlsx', '')
        evaluate(primary_filename, key, parsed_args.stage2, f'final-{file_name}')
        count += 1


if __name__ == "__main__":
    main()
